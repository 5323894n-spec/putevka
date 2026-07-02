from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Side, Border

from .models import Duty, Waybill


thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def workbook_response_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def duties_workbook(duties: list[Duty]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Наряд"
    headers = ["Дата", "Маршрут", "Выпуск", "Смена", "ТС", "Гаражный", "Водитель", "Табельный", "Выезд", "Возврат", "Пробег", "Рейсы", "Примечание"]
    ws.append(headers)
    for duty in duties:
        ws.append([
            duty.duty_date.isoformat(),
            duty.route.number,
            duty.run.number,
            duty.shift,
            duty.vehicle.plate_no,
            duty.vehicle.garage_no,
            duty.driver.full_name,
            duty.driver.personnel_no,
            duty.planned_out_time.strftime("%H:%M"),
            duty.planned_in_time.strftime("%H:%M"),
            duty.planned_mileage,
            duty.planned_trips,
            duty.note or "",
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return workbook_response_bytes(wb)


def waybills_workbook(waybills: list[Waybill]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал"
    headers = ["Номер", "Дата выписки", "Дата работы", "Водитель", "Автобус", "Госномер", "Маршрут", "Выпуск", "Выезд", "Заезд", "Пробег", "Топливо выезд", "Топливо заезд", "Статус", "Ответственный"]
    ws.append(headers)
    for item in waybills:
        ws.append([
            item.number,
            item.issue_date.isoformat(),
            item.work_date.isoformat(),
            item.driver.full_name,
            f"{item.vehicle.brand} {item.vehicle.model or ''}".strip(),
            item.vehicle.plate_no,
            item.route.number,
            item.run.number,
            item.planned_out_time.strftime("%H:%M"),
            item.planned_in_time.strftime("%H:%M"),
            item.mileage,
            item.fuel_out,
            item.fuel_in,
            item.status.value,
            item.responsible or "",
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return workbook_response_bytes(wb)


def single_waybill_workbook(waybill: Waybill) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Путевой лист"
    ws["A1"] = "ПУТЕВОЙ ЛИСТ АВТОБУСА"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"№ {waybill.number}"
    ws["A3"] = f"Организация: {waybill.organization}"
    ws["A4"] = f"Период: {waybill.valid_from.isoformat()} - {waybill.valid_to.isoformat()}"
    rows = [
        ("Марка автобуса", f"{waybill.vehicle.brand} {waybill.vehicle.model or ''}".strip()),
        ("Государственный номер", waybill.vehicle.plate_no),
        ("Гаражный номер", waybill.vehicle.garage_no),
        ("Водитель", waybill.driver.full_name),
        ("Табельный номер", waybill.driver.personnel_no),
        ("Водительское удостоверение", waybill.driver.license_no or ""),
        ("Маршрут", f"{waybill.route.number} {waybill.route.name}"),
        ("Выпуск", waybill.run.number),
        ("Время выезда", waybill.planned_out_time.strftime("%H:%M")),
        ("Время заезда", waybill.planned_in_time.strftime("%H:%M")),
        ("Одометр выезд", waybill.odometer_out),
        ("Одометр заезд", waybill.odometer_in),
        ("Пробег", waybill.mileage),
        ("Остаток топлива выезд", waybill.fuel_out),
        ("Выдано топлива", waybill.fuel_issued),
        ("Остаток топлива заезд", waybill.fuel_in),
        ("Расход по норме", waybill.norm_fuel),
        ("Фактический расход", waybill.fact_fuel),
        ("Медицинский осмотр", waybill.medical_check),
        ("Технический контроль", waybill.technical_check),
    ]
    for index, (label, value) in enumerate(rows, start=6):
        ws.cell(index, 1, label)
        ws.cell(index, 2, value)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 64
    return workbook_response_bytes(wb)

